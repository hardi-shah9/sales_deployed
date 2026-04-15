"""
Utility functions for Dreamgirl CRM System
"""
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os


def generate_excel_report(sales_data, filename='sales_report.xlsx'):
    """
    Generate Excel report from sales data with all commission details
    
    Args:
        sales_data: List of sale dictionaries
        filename: Output filename
    
    Returns:
        Path to generated Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Report"
    
    # Define headers
    headers = [
        'Date', 'Bill No', 'Salesman Name', 'Sales Amount', 
        'Discount Amount', 'Net Amount', '1% Commission', 'Discount %',
        'CD Bonus Eligible', 'CD Bonus Amount', 
        'LT Bonus', 'Total Commission'
    ]
    
    # Style definitions
    header_fill = PatternFill(start_color="8B4513", end_color="8B4513", fill_type="solid")  # Brown
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Write data
    for row_num, sale in enumerate(sales_data, 2):
        ws.cell(row=row_num, column=1, value=sale['date']).border = border
        ws.cell(row=row_num, column=2, value=sale['bill_no']).border = border
        ws.cell(row=row_num, column=3, value=sale['salesman_name']).border = border
        ws.cell(row=row_num, column=4, value=sale['sales_amt']).border = border
        ws.cell(row=row_num, column=5, value=sale['discount_amt']).border = border
        ws.cell(row=row_num, column=6, value=sale['net_amount']).border = border
        ws.cell(row=row_num, column=7, value=sale['commission_1_percent']).border = border
        ws.cell(row=row_num, column=8, value=sale['discount_percentage']).border = border
        ws.cell(row=row_num, column=9, value='YES' if sale['cd_bonus_eligible'] else 'NO').border = border
        ws.cell(row=row_num, column=10, value=sale['cd_bonus_amount']).border = border
        ws.cell(row=row_num, column=11, value=sale['lt_bonus']).border = border
        ws.cell(row=row_num, column=12, value=sale['total_commission']).border = border
    
    # Add summary row if data exists
    if sales_data:
        summary_row = len(sales_data) + 3
        ws.cell(row=summary_row, column=1, value="TOTAL").font = Font(bold=True)
        
        # Calculate totals
        total_sales = sum(s['sales_amt'] for s in sales_data)
        total_discount = sum(s['discount_amt'] for s in sales_data)
        total_net = sum(s['net_amount'] for s in sales_data)
        total_commission_1 = sum(s['commission_1_percent'] for s in sales_data)
        total_cd_bonus = sum(s['cd_bonus_amount'] for s in sales_data)
        total_lt = sum(s['lt_bonus'] for s in sales_data)
        total_commission = sum(s['total_commission'] for s in sales_data)
        
        ws.cell(row=summary_row, column=4, value=total_sales).font = Font(bold=True)
        ws.cell(row=summary_row, column=5, value=total_discount).font = Font(bold=True)
        ws.cell(row=summary_row, column=6, value=total_net).font = Font(bold=True)
        ws.cell(row=summary_row, column=7, value=total_commission_1).font = Font(bold=True)
        ws.cell(row=summary_row, column=10, value=total_cd_bonus).font = Font(bold=True)
        ws.cell(row=summary_row, column=11, value=total_lt).font = Font(bold=True)
        ws.cell(row=summary_row, column=12, value=total_commission).font = Font(bold=True)
    
    # Adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Save file
    filepath = os.path.join(os.path.dirname(__file__), filename)
    wb.save(filepath)
    return filepath


def get_salesman_summary(sales_data):
    """
    Generate salesman-wise summary from sales data
    
    Args:
        sales_data: List of sale dictionaries
    
    Returns:
        List of summary dictionaries
    """
    summary = {}
    
    for sale in sales_data:
        name = sale['salesman_name']
        if name not in summary:
            summary[name] = {
                'salesman_name': name,
                'total_sales': 0,
                'total_discount': 0,
                'total_net': 0,
                'total_commission': 0,
                'transaction_count': 0,
                'days_worked': set()
            }
        
        summary[name]['total_sales'] += sale['sales_amt']
        summary[name]['total_discount'] += sale['discount_amt']
        summary[name]['total_net'] += sale['net_amount']
        summary[name]['total_commission'] += sale['total_commission']
        summary[name]['transaction_count'] += 1
        summary[name]['days_worked'].add(sale['date'])
    
    # Convert to list and format
    result = []
    for name, data in summary.items():
        result.append({
            'salesman_name': data['salesman_name'],
            'total_sales': round(data['total_sales'], 2),
            'total_discount': round(data['total_discount'], 2),
            'total_net': round(data['total_net'], 2),
            'total_commission': round(data['total_commission'], 2),
            'transaction_count': data['transaction_count'],
            'days_worked': len(data['days_worked'])
        })
    
    return sorted(result, key=lambda x: x['total_sales'], reverse=True)
